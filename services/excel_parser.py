import openpyxl
from collections import defaultdict
from utils.logging import debug_log

def extract_instruction_blocks(xlsx_path):
	debug_log("[extract_instruction_blocks] Entered")
	wb = openpyxl.load_workbook(xlsx_path, data_only=True)
	step_blocks = []

	for sheet_name in wb.sheetnames:
		debug_log(f"📄 Reading sheet: {sheet_name}")
		ws = wb[sheet_name]
		temp_block = {"sheet": sheet_name, "step_id": None, "action": "", "steps": [], "expected": ""}
		for row in ws.iter_rows(values_only=True):
			debug_log(f"📊 Raw row: {row}")
			if all(cell is None for cell in row):
				continue
			if any(isinstance(cell, str) and cell.strip().startswith("<?xml") for cell in row):
				debug_log("⛔ Skipping XML-like row")
				continue

			id_col, action_col, nav_col, expected_col = (list(row) + [None]*4)[:4]

			if id_col or action_col:
				# save previous block if valid and not a header
				if temp_block["step_id"] and temp_block["step_id"].lower() != "step ref":
					step_blocks.append(temp_block)

				temp_block = {
					"sheet": sheet_name,
					"step_id": id_col.strip() if id_col else None,
					"action": action_col.strip() if action_col else "",
					"steps": [nav_col.strip()] if nav_col else [],
					"expected": expected_col.strip() if expected_col else ""
				}
			else:
				if nav_col:
					temp_block["steps"].append(nav_col.strip())
				if expected_col:
					temp_block["expected"] = expected_col.strip()

		# save final block if valid and not a header
		if temp_block["step_id"] and temp_block["step_id"].lower() != "step ref":
			step_blocks.append(temp_block)

	# ✅ Group by sheet
	grouped_blocks = defaultdict(list)
	for block in step_blocks:
		grouped_blocks[block["sheet"]].append(block)

	debug_log(f"[extract_instruction_blocks] Extracted {len(grouped_blocks)} sheets")
	debug_log("[extract_instruction_blocks] Exited")
	return dict(grouped_blocks)
