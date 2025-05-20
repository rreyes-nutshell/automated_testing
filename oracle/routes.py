import os
import tempfile
from flask import Blueprint, render_template, request
from utils.logging import debug_log
<<<<<<< HEAD
from openpyxl import load_workbook
from oracle.use_cases import generate_test_cases_by_sheet

oracle_runner_bp = Blueprint("oracle_runner_bp", __name__)

def extract_lines_by_sheet(xlsx_path):
    from collections import defaultdict
    wb = load_workbook(xlsx_path, data_only=True)
    lines_by_sheet = defaultdict(list)

    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for row in ws.iter_rows():
            values = [str(cell.value).strip() for cell in row if cell.value and str(cell.value).strip()]
            if values:
                lines_by_sheet[sheet].append(" | ".join(values))
    return lines_by_sheet

@oracle_runner_bp.route("/", methods=["GET", "POST"])
def index():
    debug_log("Entered")
    grouped_result = {}
    total_test_cases = 0
    sheet_count = 0

    if request.method == "POST":
        file = request.files.get("test_file")
        if file and file.filename.endswith(".xlsx"):
            with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
                file.save(tmp.name)
                sheet_lines = extract_lines_by_sheet(tmp.name)
                grouped_result, total_test_cases = generate_test_cases_by_sheet(sheet_lines)
                sheet_count = len(grouped_result)
                os.unlink(tmp.name)

    debug_log("Exited")
    return render_template("index.html", grouped_result=grouped_result, sheet_count=sheet_count, total_test_cases=total_test_cases)
=======
from oracle.use_cases import generate_test_cases_by_sheet
from services.excel_parser import extract_instruction_blocks
from ai_helpers.llm_interpreter import parse_instruction_blocks
from services.playwright_runner import run_browser_script  # âœ… Correct location

oracle_runner_bp = Blueprint("oracle_runner_bp", __name__)

@oracle_runner_bp.route("/", methods=["GET", "POST"])
async def index():  # âœ… Now async
	debug_log("Entered")
	grouped_result = {}
	total_test_cases = 0
	sheet_count = 0

	if request.method == "POST":
		file = request.files.get("test_file")
		login_url = request.form.get("login_url")
		username = request.form.get("username")
		password = request.form.get("password")
		preview_flag = request.form.get("preview") == "on"
		timeout_val = int(request.form.get("timeout", "30"))

		debug_log(f"ðŸ’Ž Received file: {file.filename if file else 'None'}")
		debug_log(f"ðŸ” Login URL: {login_url}, Username: {username}, Password: {'*' * len(password) if password else 'None'}")

		if file and file.filename.endswith(".xlsx"):
			with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
				file.save(tmp.name)
				debug_log(f"ðŸ“ Temp file saved: {tmp.name}")

				sheet_lines = extract_instruction_blocks(tmp.name)
				for sheet, blocks in sheet_lines.items():
					debug_log(f"ðŸ“„ Sheet '{sheet}' contains {len(blocks)} blocks")
					debug_log(f"ðŸ“˜ Sheet '{sheet}' - First 2 lines: {blocks[:2]}")

				grouped_result, total_test_cases = generate_test_cases_by_sheet(sheet_lines)
				sheet_count = len(grouped_result)

				step_blocks = [
					{"sheet": sheet, "steps": blocks, "timeout": timeout_val}
					for sheet, blocks in sheet_lines.items()
					if all(isinstance(block, dict) and isinstance(block.get("steps"), list) for block in blocks)
				]

				debug_log(f"âœ… Clean step blocks for LLM: {len(step_blocks)}")
				for block in step_blocks:
					debug_log(f"ðŸ’‘ LLM Step Block for Sheet '{block['sheet']}': {block['steps'][:2]}")

				llm_results = parse_instruction_blocks(step_blocks)

				for item in llm_results:
					sheet = item["sheet"]
					if sheet not in grouped_result:
						grouped_result[sheet] = {"test_case_count": 0, "content": "", "llm_steps": []}
					grouped_result[sheet].setdefault("llm_steps", []).append(item)

				# âœ… Only run one browser flow for all steps combined
				all_llm_steps = []
				for item in llm_results:
					all_llm_steps.extend(item.get("llm_json", []))

				await run_browser_script(
					all_llm_steps,
					login_url=login_url,
					username=username,
					password=password,
					preview_mode=preview_flag
				)

				os.unlink(tmp.name)

	debug_log("Exited")
	return render_template("index.html", grouped_result=grouped_result, sheet_count=sheet_count, total_test_cases=total_test_cases)
>>>>>>> 61b9b8f (Patch UI Mapper: crawler, extractor, schema updates, HTML fixes)
